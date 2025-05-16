import json
import xml.etree.ElementTree as ET
import numpy as np
from tqdm import tqdm # type: ignore
import pickle

from tifffile import TiffFile # type: ignore
from tifffile.tiffcomment import tiffcomment
from xml.etree import ElementTree
import cv2 # type: ignore
import SimpleITK as sitk # type: ignore

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from collections import Counter

from src.utils.cropping import ImageCropperApp
from src.utils.manual_alignment import ImageManualAlignmentApp
import tkinter as tk
import matplotlib.pyplot as plt
import os
from src.alignment import get_annotations
from shapely.geometry import Polygon
import copy

import os
import fnmatch


def save_downscaled_images(params_file_path, excluded_ids, brightness_factor):
    ## Get the parameters for alignment
    scans_paths, annotations_paths, annotations_names_empty, annotations_names_artefacts, annotations_names_AnalysisArea, common_ids, panels_all, output_path = get_parameters(params_file_path, excluded_ids)
    ## Get the name of the panels for each alignment (with reference panel at the end)
    data_dict = {}
    panel_alignment_dict = panels_name_alignment(panels_all, scans_paths)

    unique_name_alignments = set()
    root = tk.Tk()
    for id in tqdm(common_ids, desc="Loading downscaled images", unit="Patient"):
        data_dict[id] = {}  # Initialize sub-dictionary for this patient
        for name_alignment, scans_alignment_paths in panel_alignment_dict.items():
            panels = name_alignment.split('_')
            # Add the name_alignment to the set (sets automatically handle uniqueness)
            unique_name_alignments.add(name_alignment)
            scans_path1, scans_path2 = [path for path in scans_alignment_paths]
            # Load imgs downscaled and metadata
            (channels1, scale_percent1, channels2, scale_percent2,  
            img1, img2, img1_resize, img2_resize, img1_resize_ori, img2_resize_ori) = load_downscaled_imgs(scans_path1, scans_path2, id)

            if annotations_paths:
                annotations_resized = []
                for panel in panels:
                    # Get the panel path
                    index = panels_all.index(panel)
                    annotations = annotations_paths[index]
                    # Get the file path corresponding to id
                    # Get annotations
                    artefacts_empty_alignment = {}
                    analysis_area_alignment = {}
                    # geojson_file_path = next((os.path.join(annotations, f) for f in os.listdir(annotations) if id in f and f.endswith(".geojson") and not f.startswith(".")), None)
                    annotation_file_path = next(
                        (os.path.join(annotations, f) for f in os.listdir(annotations) 
                        if id in f and (f.endswith(".annotations") or f.endswith(".geojson")) and not f.startswith(".")), 
                        None
                    )
                    ## Get the annotations
                    artefacts_empty_alignment, analysis_area_alignment = get_annotations(annotation_file_path, panel, artefacts_empty_alignment, analysis_area_alignment, annotations_names_empty, annotations_names_artefacts, annotations_names_AnalysisArea)
                    analysis_area_gdf_poly = analysis_area_alignment[panel]
                    if index == 0:
                        analysis_area_polygons = get_polygon(analysis_area_gdf_poly, scale_percent1, image_shape = img1_resize.shape, img_resize = img1_resize)
                        annotations_resized.append(analysis_area_polygons)
                    else:
                        analysis_area_polygons = get_polygon(analysis_area_gdf_poly, scale_percent2, image_shape = img2_resize.shape, img_resize = img2_resize)
                        annotations_resized.append(analysis_area_polygons)
                    

            #plt.tight_layout()
            ## Cropping step
            if annotations_paths:
                app = ImageCropperApp(tk.Toplevel(root), img1_resize, img2_resize, panels, brightness_factor, annotations_resized)
            else:
                app = ImageCropperApp(tk.Toplevel(root), img1_resize, img2_resize, panels, brightness_factor)

            root.mainloop()
            if app.saved:
                cropped_images_dict = app.cropped_images
                if panels[0] in cropped_images_dict:
                    img1_resize = cropped_images_dict[panels[0]][0]
                    # Convert to sitk image format
                    img1 = sitk.GetImageFromArray(img1_resize)
                    # Get the coordinates of the cropping for origin coordinate/ top-left corner of the image
                    crop_coords1 = cropped_images_dict[panels[0]][1]
                else:
                    crop_coords1 = (0, 0)
                if panels[1] in cropped_images_dict:
                    img2_resize = cropped_images_dict[panels[1]][0]
                    # Convert to sitk image format
                    img2 = sitk.GetImageFromArray(img2_resize)
                    # Get the coordinates of the cropping for origin coordinate/ top-left corner of the image
                    crop_coords2 = cropped_images_dict[panels[1]][1]
                else:
                    crop_coords2 = (0, 0)
            else:
                crop_coords1 = (0, 0)
                crop_coords2 = (0, 0)

            ## Manual alignment step
            app = ImageManualAlignmentApp(tk.Toplevel(root), img1_resize, img2_resize, panels, brightness_factor)
            root.mainloop()
            
            
            if hasattr(app, 'manually_aligned_image1'):
                img1_resize = copy.deepcopy(app.manually_aligned_image1)
                # Convert to sitk image format
                img1 = sitk.GetImageFromArray(img1_resize)
                img1 = sitk.Cast(img1, sitk.sitkFloat32)
                # Get the coordinates of the cropping for origin coordinate/ top-left corner of the image
                manual_alignment_rotation_angle = app.angle
                manual_alignment_rotation_offset = app.offset
                manual_alignment_rotation_shape = app.orig_shape
                manual_alignment_displacement = (app.trans_x, app.trans_y)
            else:
                manual_alignment_rotation_angle = 0
                manual_alignment_rotation_offset = (0, 0)
                manual_alignment_rotation_shape = (0, 0)
                manual_alignment_displacement = (0, 0)

            '''
            # Plotting the blue image with alpha blending and then overlaying the red image
            plt.figure(figsize=(10, 5))
            # Plot the first image (blue) with alpha blending
            plt.imshow(img1_resize, cmap='Blues', alpha=0.6)  # Adjust alpha for blending
            # Plot the second image (red) on top of the blue image
            plt.imshow(img2_resize, cmap='Reds', alpha=0.4)  # Red image on top with a different alpha for blending
            plt.title('Overlay of Blue and Red Images')
            plt.axis('off')

            #plt.tight_layout()
            plt.savefig('/Users/leohermet/Downloads/overlay_blue_red_image_test.png')  # Save as PNG with tight bounding box
            '''

            
            '''
            if app.saved:
                cropped_images_dict = app.cropped_images
                img1_resize = cropped_images_dict[panels[0]][0]
                img2_resize = cropped_images_dict[panels[1]][0]
                # Convert to sitk image format
                img1 = sitk.GetImageFromArray(img1_resize)
                img2 = sitk.GetImageFromArray(img2_resize)
                # Get the coordinates of the cropping for origin coordinate/ top-left corner of the image
                crop_coords1 = cropped_images_dict[panels[0]][1]
                crop_coords2 = cropped_images_dict[panels[1]][1]
            else:
                crop_coords1 = (0, 0)
                crop_coords2 = (0, 0)
            '''

            data_dict[id][name_alignment] = {
                "channels1": channels1,
                "scale_percent1": scale_percent1,
                "channels2": channels2,
                "scale_percent2": scale_percent2,
                "img1": img1,
                "img2": img2,
                "img1_resize": img1_resize,
                "img2_resize": img2_resize,
                "crop_coords1": crop_coords1,
                "crop_coords2": crop_coords2,
                "manual_alignment_displacement": manual_alignment_displacement,
                "manual_alignment_rotation_angle": manual_alignment_rotation_angle,
                "manual_alignment_rotation_shape": manual_alignment_rotation_shape,
                "img1_resize_ori": img1_resize_ori,
                "img2_resize_ori": img2_resize_ori,
                "img1_shape_ori": img1_resize_ori.shape,
                "img2_shape_ori": img2_resize_ori.shape,
            }
    root.destroy()
    # Add params
    data_dict["params"] = {"annotations_paths":annotations_paths, "annotations_names_empty":annotations_names_empty, "annotations_names_artefacts":annotations_names_artefacts, "annotations_names_AnalysisArea":annotations_names_AnalysisArea, "common_ids" : common_ids, "panels" : panels_all, "output_path" : output_path}
    # Save the dictionary to a file
    with open(output_path + "downscaled_images.pkl", "wb") as file:
        pickle.dump(data_dict, file)
    print(f'Data saved to {output_path + "downscaled_images.pkl"}')

    ## Create DataFrame for common_ids with empty status columns for manual input
    # Initialize the data dictionary with 'Sample id' and 'Excluded' columns
    data = {
        'Sample id': common_ids,
        'Excluded': ['No'] * len(common_ids)
    }
    # Create a column for each unique name_alignment and add it to the data dictionary
    for name_alignment in unique_name_alignments:
        # Dynamically create a new column name for each unique 'name_alignment'
        column_name = f'Status {name_alignment} alignment'
        data[column_name] = [None] * len(common_ids)  # Initialize with None for manual input
    # Add the 'Comments' column after all the Status columns
    data['Comments'] = [None] * len(common_ids)
    df = pd.DataFrame(data)

    ## Create DataFrame for excluded_ids with empty status columns for manual input
    excluded_data = {
        'Sample id': excluded_ids,
        'Excluded': ['Yes']*len(excluded_ids)
    }
    # Create a column for each unique name_alignment and add it to the data dictionary
    for name_alignment in unique_name_alignments:
        # Dynamically create a new column name for each unique 'name_alignment'
        column_name = f'Status {name_alignment} alignment'
        excluded_data[column_name] = [None] * len(excluded_ids)  # Initialize with None for manual input
    # Add the 'Comments' column after all the Status columns
    excluded_data['Comments'] = [None] * len(excluded_ids)

    # Concatenate the common_ids DataFrame with the excluded_data DataFrame at the bottom
    df = pd.concat([df, pd.DataFrame(excluded_data)], ignore_index=True)

    df['Sample id'] = '"' + df['Sample id'].astype(str) + '"'

    # Save to Excel (instead of CSV) to enable styling
    excel_file_path = output_path + "Alignments_validated.xlsx"
    df.to_excel(excel_file_path, index=False, engine='openpyxl')

    # Load the Excel workbook and sheet
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Define fill colors for gray, light green, and light red
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    light_green_fill = PatternFill(start_color="A8E6A1", end_color="A8E6A1", fill_type="solid")  # Light green
    light_red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red

    # Apply gray fill to all cells in the row if 'Excluded' is 'Yes'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):  # Columns 1-5 are all the columns (Sample id, Excluded, Status, Comments)
        if row[1].value == 'Yes':  # Check if 'Excluded' (2nd column) is 'Yes'
            for cell in row:  # Apply the gray fill to all cells in the row
                cell.fill = gray_fill

    # Create conditional formatting for 'Status first alignment' and 'Status second alignment' columns
    # Light green for 'Y'
    green_rule = FormulaRule(formula=['$C2="Y"'], fill=light_green_fill)  # Column C = 'Status first alignment'
    ws.conditional_formatting.add('C2:C{}'.format(ws.max_row), green_rule)

    green_rule_2 = FormulaRule(formula=['$D2="Y"'], fill=light_green_fill)  # Column D = 'Status second alignment'
    ws.conditional_formatting.add('D2:D{}'.format(ws.max_row), green_rule_2)

    # Light red for 'N'
    red_rule = FormulaRule(formula=['$C2="N"'], fill=light_red_fill)  # Column C = 'Status first alignment'
    ws.conditional_formatting.add('C2:C{}'.format(ws.max_row), red_rule)

    red_rule_2 = FormulaRule(formula=['$D2="N"'], fill=light_red_fill)  # Column D = 'Status second alignment'
    ws.conditional_formatting.add('D2:D{}'.format(ws.max_row), red_rule_2)

    # Adjust the column widths for better readability of the headers
    ws.column_dimensions['C'].width = 25  # Status first alignment
    ws.column_dimensions['D'].width = 25  # Status second alignment
    ws.column_dimensions['E'].width = 60  # Status second alignment

    # Save the modified workbook with conditional formatting and adjusted column widths
    wb.save(excel_file_path)

    print(f'Data saved to {excel_file_path}')









        
def get_parameters(params_file_path, excluded_ids):
    ''' Get the parameters define in the visualization step. '''
    # Load parameters
    with open(params_file_path, "r") as infile:
        params_dict = json.load(infile)
    scans_paths = params_dict.get("scans_paths")
    annotations_paths = params_dict.get("annotations_paths")
    annotations_names_AnalysisArea = params_dict.get("annotations_names_AnalysisArea")
    annotations_names_empty = params_dict.get("annotations_names_empty")
    annotations_names_artefacts = params_dict.get("annotations_names_artefacts")
    common_ids = params_dict.get("common_ids")
    panels = params_dict.get("panels")
    output_path = params_dict.get("output_path")

    ## Exclude specific ids
    common_ids = [id for id in common_ids if id not in excluded_ids]
    
    return scans_paths, annotations_paths, annotations_names_empty, annotations_names_artefacts, annotations_names_AnalysisArea, common_ids, panels, output_path

def panels_name_alignment(panels, folder_paths):
    ''' Get the name of the panels and corresponding folder for each alignment. '''
    # If there is 2 panels in total then the reference panel is at the end
    if len(panels) == 2:
        panel_alignment_dict = {"_".join(panels): folder_paths}
    # If there is 3 panels in total then the reference panel is in the middle
    elif len(panels) == 3:
        panel_alignment_dict = {
            "_".join([panels[0], panels[1]]): [folder_paths[0], folder_paths[1]],
            "_".join([panels[2], panels[1]]): [folder_paths[2], folder_paths[1]],
        }
    
    return panel_alignment_dict

def load_downscaled_imgs(folder_path1, folder_path2, id):
    # Define path of the qptiff or tiff
    path1 = "\n".join([os.path.join(folder_path1, f) for f in os.listdir(folder_path1) if fnmatch.fnmatch(f.lower(), f"*{id}*.qptiff") or fnmatch.fnmatch(f.lower(), f"*{id}*.tif")])
    path2 = "\n".join([os.path.join(folder_path2, f) for f in os.listdir(folder_path2) if fnmatch.fnmatch(f.lower(), f"*{id}*.qptiff") or fnmatch.fnmatch(f.lower(), f"*{id}*.tif")])

    #path1 = folder_path1 + id + f"_panel_{panels[0]}.unmixed.qptiff"
    #path2 = folder_path2 + id + f"_panel_{panels[1]}.unmixed.qptiff"
    # print("path1:", path1)
    # print("path2:", path2)

    if path1.endswith('.qptiff') and path2.endswith('.qptiff'):
        ## Load images
        channels1, img1_resize, sizeY_compressed1, sizeY_fullres1 = get_data_alignment_qptiff(path1)
        channels2, img2_resize, sizeY_compressed2, sizeY_fullres2 = get_data_alignment_qptiff(path2)
    
    elif path1.endswith('.tif') and path2.endswith('.tif'):
        channels1, img1_resize, sizeY_compressed1, sizeY_fullres1 = get_data_alignment_tif(path1)
        channels2, img2_resize, sizeY_compressed2, sizeY_fullres2 = get_data_alignment_tif(path2)

    scale_percent1 = sizeY_compressed1/sizeY_fullres1
    scale_percent2 = sizeY_compressed2/sizeY_fullres2

    img1_resize, scale_percent1, img2_resize, scale_percent2 = get_same_compression(id, path1, img1_resize, scale_percent1, sizeY_fullres1, path2, img2_resize, scale_percent2, sizeY_fullres2)
    

    # Scale images to 8bit
    img1_8bit = cv2.convertScaleAbs(img1_resize)
    img2_8bit = cv2.convertScaleAbs(img2_resize)

    img1_resize_ori = copy.deepcopy(img1_8bit)
    img2_resize_ori = copy.deepcopy(img2_8bit)


    # Convert to sitk image format
    img1 = sitk.GetImageFromArray(img1_8bit)
    img2 = sitk.GetImageFromArray(img2_8bit)
    img1 = sitk.Cast(img1, sitk.sitkFloat32)
    img2 = sitk.Cast(img2, sitk.sitkFloat32)
    
    # Save images
    img1_arr = sitk.GetArrayFromImage(img1)
    img2_arr = sitk.GetArrayFromImage(img2)

    return channels1, scale_percent1, channels2, scale_percent2, img1, img2, img1_arr, img2_arr, img1_resize_ori, img2_resize_ori


def get_data_alignment_qptiff(path_scan):
    tif = TiffFile(path_scan)
    tif_tags = {tag.name: tag.value for tag in tif.pages[0].tags.values()}
    # Parse the XML content
    xml_content = tif_tags.get("ImageDescription", "").replace("\r\n", "").replace("\t", "")
    channels = generate_channels_list(xml_content)
    # Load the most compressed DAPI image in .pages
    img_resize = tif.series[0].levels[-1].asarray()[0]
    sizeY_fullres = tif_tags['ImageLength']
    sizeY_compressed = img_resize.shape[0]

    return channels, img_resize, sizeY_compressed, sizeY_fullres

def extract_json_from_xml(xml_content):
    """Extracts JSON content from an XML element that contains it as text."""
    root = ET.fromstring(xml_content)
    
    # Find the first element containing JSON-like text
    json_text = None
    for elem in root.iter():
        if elem.text and "{" in elem.text and "spectra" in elem.text:
            json_text = elem.text.strip()
            break

    if not json_text:
        raise ValueError("No JSON data found in XML.")
    
    return json.loads(json_text)

def opal_to_rgb(opal_name):
    """Converts Opal fluorophores to RGB values (approximate mappings)."""
    opal_rgb_map = {
    "Opal 480": '65535',      
    "Opal 520": '65280',      
    "Opal 570": '16776960',   
    "Opal 620": '16744448',   
    "Opal 690": '16711680',   
    "Opal 780": '16777215',  
    "Sample AF": '0',       
    "DAPI": '255'              
    }
    return opal_rgb_map.get(opal_name, '0')  # Default to white

def generate_channels_list(xml_content):
    """Generates a structured list of channels with IDs and RGB values."""
    json_data = extract_json_from_xml(xml_content)
    spectra = json_data.get("spectra", [])
    bands = json_data.get("bands", [])

    channels = []
    for i, spectrum in enumerate(spectra):
        fluor = spectrum.get("fluor", "Unknown")
        marker = spectrum.get("marker", "Unknown")
        rgb = opal_to_rgb(fluor)

        channels.append({
            "id": i + 1,
            "name": f"{marker} ({fluor})",
            "rgb": rgb
        })

    return channels

def get_data_alignment_tif(path_scan):
    tif = TiffFile(path_scan, is_ome=False)
    # Load the most compressed DAPI image in .levels
    img_resize = tif.series[0].levels[-1].asarray()[0]
    xml = tiffcomment(path_scan)
    root = ElementTree.fromstring(xml.replace("\n", "").replace("\t", ""))
    ## Get the size of the full resolution image
    # Be carefull only the first value works for full res images, other don't match the downscale images sizes
    sizeY_fullres = int(list(dict.fromkeys([x.get("sizeY") for x in root.iter() if x.tag == "dimension"]))[0])
    # Get the size of the full image resolution this way if "sizeY" don't give the same sizes
    # and tif.series[0].levels[0].asarray()[0].shape[0] takes too long to compute
    #sizeY_fullres = tif.pages[0].asarray().shape[0]
    sizeY_compressed = img_resize.shape[0]
    # sizeY = list(dict.fromkeys([x.get("sizeY") for x in root.iter() if x.tag == "dimension"]))
    channels = [
        {"id": elem.get("id"), "name": elem.get("name"), "rgb": elem.get("rgb")}
        for elem in root.iter() if elem.tag == "channel"
    ]
    return channels, img_resize, sizeY_compressed, sizeY_fullres

def get_same_compression(id, path1, img1_resize, scale_percent1, sizeY_fullres1, path2, img2_resize, scale_percent2, sizeY_fullres2):
    # print("Compression")
    comp_lvl1 = round((1/scale_percent1))
    comp_lvl2 = round((1/scale_percent2))
    # print(f"img1: 1/{comp_lvl1}ème")
    # print(f"img2: 1/{comp_lvl2}ème")

    # Check if both images are the same size
    if comp_lvl2 < comp_lvl1:
        index_comp = -1
        print(f"Patient {id}: Not same compression level")
        tif = TiffFile(path1)
        while comp_lvl2 < comp_lvl1:
            index_comp-=1
            img1_resize = tif.series[0].levels[index_comp].asarray()[0]
            scale_percent1 = img1_resize.shape[0] / sizeY_fullres1
            comp_lvl1 = round((1/scale_percent1))
            print("Finding the correct compression")
            print(f"img1: 1/{comp_lvl1}ème")
            print(f"img2: 1/{comp_lvl2}ème")
    if comp_lvl1 < comp_lvl2:
        index_comp = -1
        print(f"Patient {id}: Not same compression level")
        tif = TiffFile(path2)
        while comp_lvl1 < comp_lvl2:
            index_comp-=1
            img2_resize = tif.series[0].levels[index_comp].asarray()[0]
            scale_percent2 = img2_resize.shape[0] / sizeY_fullres2
            comp_lvl2 = round((1/scale_percent2))
            print("Finding the correct compression")
            print(f"img1: 1/{comp_lvl1}ème")
            print(f"img2: 1/{comp_lvl2}ème")
    
    return img1_resize, scale_percent1, img2_resize, scale_percent2
'''
def get_same_compression(path1, img1_resize, scale_percent1, sizeY1, path2, img2_resize, scale_percent2, sizeY2):
    # print("Compression")
    comp_lvl1 = round((1/scale_percent1))
    comp_lvl2 = round((1/scale_percent2))
    # print(f"img1: 1/{comp_lvl1}ème")
    # print(f"img2: 1/{comp_lvl2}ème")
    print(sizeY1)
    print(sizeY2)
    # Check if both images are the same size
    if comp_lvl2 < comp_lvl1:
        index_comp = -1
        print(f"Patient {id}: Not same compression level")
        tif = TiffFile(path1)
        while comp_lvl2 < comp_lvl1:
            index_comp-=1
            img1_resize = tif.series[0].levels[index_comp].asarray()[0]
            scale_percent1 = int(sizeY1[index_comp]) / int(sizeY1[0])
            comp_lvl1 = round((1/scale_percent1))
            print("Finding the correct compression")
            print(f"img1: 1/{comp_lvl1}ème")
            print(f"img2: 1/{comp_lvl2}ème")
    if comp_lvl1 < comp_lvl2:
        index_comp = -1
        print(f"Patient {id}: Not same compression level")
        tif = TiffFile(path2)
        while comp_lvl1 < comp_lvl2:
            index_comp-=1
            img2_resize = tif.series[0].levels[index_comp].asarray()[0]
            scale_percent2 = int(sizeY2[index_comp]) / int(sizeY2[0])
            comp_lvl2 = round((1/scale_percent2))
            print("Finding the correct compression")
            print(f"img1: 1/{comp_lvl1}ème")
            print(f"img2: 1/{comp_lvl2}ème")
    
    return img1_resize, scale_percent1, img2_resize, scale_percent2
'''


def get_polygon(gdf, scale_percent, image_shape, img_resize = None, c = "R"):
    annotation = []
    for geometry in gdf.geometry:
        if geometry.geom_type == 'Polygon':
            annotation.append(np.array(geometry.exterior.coords))
            for hole in geometry.interiors:
                annotation.append(np.array(hole.coords))
        elif geometry.geom_type == 'MultiPolygon':
            for part in geometry:
                annotation.append(np.array(part.exterior.coords))
                for hole in part.interiors:
                    annotation.append(np.array(hole.coords))
    
    # Convert to polygon object
    polygons = []
    for array in annotation:
        scaled_points = array*scale_percent
        # Convert array to list of tuples
        points = [(x, y) for x, y in scaled_points]
        
        # Create Polygon object
        polygons.append(Polygon(points))

    return polygons


def shift_annotations(annotations, crop_coords):
    """
    Shifts the coordinates of polygons in annotations by the given crop_coords2.
    Then, plots the shifted polygons.

    Parameters:
    - annotations (list): A list where the second element is a list of polygons.
    - crop_coords (tuple): The (x, y) shift to be applied to the polygons.
    """
    shift_x, shift_y = crop_coords  # Extract the shift values
    shifted_polygons = []

    # Iterate through each polygon in annotations
    for polygon in annotations:
        # Shift all points in the polygon by the crop_coords2 shift
        shifted_polygon = [(x - shift_x, y - shift_y) for x, y in zip(polygon.exterior.xy[0], polygon.exterior.xy[1])]
        # Create a new Polygon with the shifted coordinates
        shifted_polygon = Polygon(shifted_polygon)
        shifted_polygons.append(shifted_polygon)

    return shifted_polygons

